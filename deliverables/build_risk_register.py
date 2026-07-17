"""
build_risk_register.py
Creates deliverables/risk_register.xlsx — likelihood x impact = score (formula-
driven, not hardcoded), color-scaled for at-a-glance triage. Covers project-level
risks (this portfolio project's own methodology limitations) plus the top-ranked
opportunity's implementation risks, since both are "risk registers" a reviewer
would reasonably expect from this deliverable set.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Risk Register"

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F6FEB")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
LEGEND_FILL = PatternFill("solid", fgColor="FFF2CC")
thin = Side(style="thin", color="D0D5DD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

COLUMNS = [
    ("Risk ID", 10), ("Category", 16), ("Risk Description", 42), ("Likelihood (1-5)", 14),
    ("Impact (1-5)", 12), ("Risk Score", 12), ("Owner / Stakeholder", 20),
    ("Mitigation", 44), ("Status", 14),
]

for i, (title, width) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=1, column=i, value=title)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[1].height = 30

# Likelihood/Impact are hardcoded judgment inputs (blue, per convention);
# Risk Score is a formula (black) = Likelihood * Impact, never hardcoded.
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")

ROWS = [
    # (id, category, description, likelihood, impact, owner, mitigation, status)
    ("R01", "Project Methodology", "Dataset ratings are mostly ESTIMATED/ASSUMPTION rather than primary industry data, since this is a public, non-proprietary research project.", 4, 3, "Project author", "Explicit VERIFIED/ESTIMATED/ASSUMPTION tagging on every field; sensitivity analysis shows the top ranking is not fragile to this uncertainty.", "Accepted / Disclosed"),
    ("R02", "Project Methodology", "NLP research corpus covers 8 of 14 technologies, not all 14, because bulk literature-API access (OpenAlex, arXiv) was blocked by the build environment's network policy.", 5, 2, "Project author", "Corpus and method documented as extensible; openalex_pull_template.py shows the exact code to run in an open-network environment.", "Accepted / Disclosed"),
    ("R03", "Project Methodology", "Scoring model min-max normalization means scores are relative to this specific 14-technology set, not an absolute investment-grade rating.", 5, 2, "Project author", "Stated explicitly in white paper and data dictionary as a modeling limitation, not hidden.", "Accepted / Disclosed"),
    ("R04", "Project Methodology", "Power BI Desktop (.pbix) could not be produced directly, since it is Windows-only desktop software unavailable in this build environment.", 5, 2, "Project author", "Delivered a full Power BI-ready data model, a page-by-page build guide, and a working interactive HTML dashboard preview as a substitute.", "Mitigated"),
    ("R05", "Data / IP", "Some illustrative partnership organizations and cost figures could be mistaken for confirmed relationships or Kinectrics-sourced numbers.", 2, 4, "Project author", "Every illustrative figure and partner name is explicitly labeled 'illustrative' / 'not confirmed' throughout every deliverable.", "Mitigated"),
    ("R06", "Implementation (Top Opportunity)", "Predictive-maintenance pilot may lack sufficient historical labeled failure data for the chosen equipment class.", 3, 5, "Reliability/Asset Mgmt team", "Make data-access confirmation the explicit Phase 1 stage-gate before any model-development spend (see business case, Section 13).", "Open"),
    ("R07", "Implementation (Top Opportunity)", "Low model explainability could erode plant engineers' trust and adoption of maintenance recommendations.", 3, 4, "Plant maintenance engineers", "Favor interpretable model classes (gradient-boosted trees, survival analysis) over black-box deep learning; run in advisory/shadow mode first.", "Open"),
    ("R08", "Implementation (Top Opportunity)", "New data pipeline connecting OT sensor networks to an analytics environment may be blocked or delayed by security review.", 3, 3, "IT/OT security team", "Engage IT/OT security in Phase 1 planning, not at deployment.", "Open"),
    ("R09", "Implementation (Top Opportunity)", "Over-alerting on low-confidence predictions causes engineers to ignore the tool ('alarm fatigue').", 3, 3, "Plant maintenance engineers", "Set a conservative confidence threshold for the Phase 3 pilot; tune based on shadow-mode feedback before any wider rollout.", "Open"),
    ("R10", "Implementation (Top Opportunity)", "Scope creep — expanding to additional equipment classes before the pilot proves value on the first one.", 2, 3, "Innovation Hub leadership", "Roadmap's stage-gate structure explicitly blocks expansion until Phase 3 success criteria are met.", "Open"),
]

start_row = 2
for r, row in enumerate(ROWS, start=start_row):
    rid, cat, desc, like, impact, owner, mitig, status = row
    ws.cell(row=r, column=1, value=rid).font = BLACK
    ws.cell(row=r, column=2, value=cat).font = BLACK
    c3 = ws.cell(row=r, column=3, value=desc); c3.font = BLACK; c3.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=4, value=like).font = BLUE
    ws.cell(row=r, column=5, value=impact).font = BLUE
    score_cell = ws.cell(row=r, column=6, value=f"=D{r}*E{r}")
    score_cell.font = BLACK
    ws.cell(row=r, column=7, value=owner).font = BLACK
    c8 = ws.cell(row=r, column=8, value=mitig); c8.font = BLACK; c8.alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=9, value=status).font = BLACK
    for col in range(1, 10):
        ws.cell(row=r, column=col).border = BORDER
    ws.row_dimensions[r].height = 46

last_row = start_row + len(ROWS) - 1

# Color scale on Risk Score column (green=low, red=high)
rule = ColorScaleRule(
    start_type="min", start_color="63BE7B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="F8696B",
)
ws.conditional_formatting.add(f"F{start_row}:F{last_row}", rule)

# Table for filter/sort
tab = Table(displayName="RiskRegister", ref=f"A1:I{last_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)

# Legend sheet
legend = wb.create_sheet("Legend")
legend["A1"] = "Risk Register — Legend & Method"
legend["A1"].font = Font(name=FONT, bold=True, size=13)
legend_lines = [
    "", "Likelihood and Impact are author judgment inputs, each rated 1 (lowest) to 5 (highest).",
    "Risk Score = Likelihood × Impact (formula in column F; recalculates if you change the blue input cells).",
    "Color scale: green = lower risk score, red = higher risk score.",
    "", "Categories:",
    "  Project Methodology — limitations of this portfolio project's own research process, disclosed for transparency.",
    "  Data / IP — risks around how figures/partnerships in this project could be misread as confirmed fact.",
    "  Implementation (Top Opportunity) — real-world risks to executing the #1-ranked recommendation (AI/ML predictive maintenance) if a sponsor chose to proceed.",
    "", "This is an independent portfolio project and is not affiliated with, commissioned by, or reviewed by Kinectrics Inc.",
]
for i, line in enumerate(legend_lines, start=2):
    cell = legend.cell(row=i, column=1, value=line)
    cell.font = Font(name=FONT, size=10, italic=(line.startswith("  ") or line == ""))
legend.column_dimensions["A"].width = 110

wb.save("deliverables/risk_register.xlsx")
print("Wrote deliverables/risk_register.xlsx")
